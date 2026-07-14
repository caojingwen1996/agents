from datetime import date

import openpyxl
import pytest

from grid_dashboard.errors import WorkbookValidationError
from grid_dashboard.excel_io import load_workbook


HEADERS = [
    "日期",
    "类型",
    "网格编号",
    "数量",
    "成交价",
    "手续费",
    "印花税",
    "现金金额",
    "备注",
]


def write_book(path, *, sell_quantity=100, sell_grid="G3"):
    book = openpyxl.Workbook()
    config = book.active
    config.title = "配置"
    config.append(["字段", "值"])
    config.append(["股票代码", "000001"])
    config.append(["图表标题", "网格提款足迹"])
    config.append(["行情复权", "不复权"])
    config.append(["策略预算", 100000])

    trades = book.create_sheet("交易流水")
    trades.append(HEADERS)
    trades.append([date(2025, 1, 2), "买入", "G3", 100, 10, 5, 0, 0, "建仓"])
    trades.append(
        [date(2025, 1, 10), "卖出", sell_grid, sell_quantity, 11, 5, 1.1, 0, "止盈"]
    )
    book.save(path)


def test_load_workbook_preserves_code_and_row_order(tmp_path):
    path = tmp_path / "交易记录.xlsx"
    write_book(path)

    loaded = load_workbook(path)

    assert loaded.settings.stock_code == "000001"
    assert loaded.settings.strategy_budget == 100000
    assert [trade.excel_row for trade in loaded.trades] == [2, 3]
    assert loaded.trades[0].grid_id == "G3"


def test_sell_without_grid_reports_sheet_row_and_field(tmp_path):
    path = tmp_path / "交易记录.xlsx"
    write_book(path, sell_grid="")

    with pytest.raises(WorkbookValidationError) as error:
        load_workbook(path)

    assert "交易流水第 3 行" in str(error.value)
    assert "网格编号" in str(error.value)


def test_missing_required_sheet_is_rejected(tmp_path):
    path = tmp_path / "交易记录.xlsx"
    book = openpyxl.Workbook()
    book.active.title = "配置"
    book.save(path)

    with pytest.raises(WorkbookValidationError, match="缺少工作表：交易流水"):
        load_workbook(path)


def test_unknown_trade_type_names_row_and_field(tmp_path):
    path = tmp_path / "交易记录.xlsx"
    write_book(path)
    book = openpyxl.load_workbook(path)
    book["交易流水"]["B3"] = "撤单"
    book.save(path)

    with pytest.raises(WorkbookValidationError, match="交易流水第 3 行：类型"):
        load_workbook(path)


def test_workbook_requires_at_least_one_buy(tmp_path):
    path = tmp_path / "交易记录.xlsx"
    write_book(path)
    book = openpyxl.load_workbook(path)
    sheet = book["交易流水"]
    sheet.delete_rows(2, 1)
    book.save(path)

    with pytest.raises(WorkbookValidationError, match="至少需要一笔买入记录"):
        load_workbook(path)
