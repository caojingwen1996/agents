from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import cast

import openpyxl

from .errors import WorkbookValidationError
from .models import Settings, Trade, TradeKind, WorkbookData

CONFIG_SHEET = "配置"
TRADES_SHEET = "交易流水"
TRADE_HEADERS = (
    "日期",
    "类型",
    "网格编号",
    "数量",
    "成交价",
    "手续费",
    "印花税",
    "现金金额",
    "备注",
)
TRADE_KINDS = {"买入", "卖出", "分红"}


def _error(sheet: str, row: int | None, field: str, reason: str) -> WorkbookValidationError:
    location = f"{sheet}第 {row} 行" if row is not None else sheet
    return WorkbookValidationError(f"{location}：{field}{reason}")


def _decimal(value, sheet: str, row: int, field: str, *, default_zero=False) -> Decimal:
    if value in (None, "") and default_zero:
        return Decimal("0")
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, ValueError, AttributeError):
        raise _error(sheet, row, field, "必须是数字") from None
    if not number.is_finite():
        raise _error(sheet, row, field, "必须是有限数字")
    return number


def _date(value, row: int) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            pass
    raise _error(TRADES_SHEET, row, "日期", "格式错误，请使用 YYYY-MM-DD")


def _stock_code(value) -> str:
    if isinstance(value, bool) or value in (None, ""):
        raise _error(CONFIG_SHEET, None, "股票代码", "不能为空")
    if isinstance(value, (int, float)) and float(value).is_integer():
        code = f"{int(value):06d}"
    else:
        code = str(value).strip()
        if code.isdigit():
            code = code.zfill(6)
    if len(code) != 6 or not code.isdigit():
        raise _error(CONFIG_SHEET, None, "股票代码", "必须是六位 A 股代码")
    return code


def _read_settings(sheet) -> Settings:
    values = {}
    for key, value, *_ in sheet.iter_rows(min_row=2, values_only=True):
        if key not in (None, ""):
            values[str(key).strip()] = value

    required = ("股票代码", "图表标题", "行情复权", "策略预算")
    for field in required:
        if field not in values or values[field] in (None, ""):
            raise _error(CONFIG_SHEET, None, field, "不能为空")

    adjustment = str(values["行情复权"]).strip()
    if adjustment != "不复权":
        raise _error(CONFIG_SHEET, None, "行情复权", "首版只支持“不复权”")
    budget = _decimal(values["策略预算"], CONFIG_SHEET, 0, "策略预算")
    if budget <= 0:
        raise _error(CONFIG_SHEET, None, "策略预算", "必须大于 0")

    return Settings(
        stock_code=_stock_code(values["股票代码"]),
        chart_title=str(values["图表标题"]).strip(),
        adjustment=adjustment,
        strategy_budget=budget,
    )


def _read_trades(sheet) -> tuple[Trade, ...]:
    headers = [cell.value for cell in sheet[1]]
    missing = [header for header in TRADE_HEADERS if header not in headers]
    if missing:
        raise WorkbookValidationError(f"{TRADES_SHEET}：缺少列：{'、'.join(missing)}")
    columns = {header: headers.index(header) for header in TRADE_HEADERS}
    trades = []

    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in values):
            continue
        kind_value = values[columns["类型"]]
        kind = "" if kind_value is None else str(kind_value).strip()
        if kind not in TRADE_KINDS:
            raise _error(TRADES_SHEET, row_number, "类型", "只允许买入、卖出、分红")

        grid_value = values[columns["网格编号"]]
        grid_id = None if grid_value in (None, "") else str(grid_value).strip()
        if kind in {"买入", "卖出"} and not grid_id:
            raise _error(TRADES_SHEET, row_number, "网格编号", "不能为空")

        quantity = _decimal(
            values[columns["数量"]], TRADES_SHEET, row_number, "数量", default_zero=True
        )
        price = _decimal(
            values[columns["成交价"]], TRADES_SHEET, row_number, "成交价", default_zero=True
        )
        commission = _decimal(
            values[columns["手续费"]], TRADES_SHEET, row_number, "手续费", default_zero=True
        )
        stamp_tax = _decimal(
            values[columns["印花税"]], TRADES_SHEET, row_number, "印花税", default_zero=True
        )
        cash_amount = _decimal(
            values[columns["现金金额"]], TRADES_SHEET, row_number, "现金金额", default_zero=True
        )
        if kind in {"买入", "卖出"} and quantity <= 0:
            raise _error(TRADES_SHEET, row_number, "数量", "必须大于 0")
        if kind in {"买入", "卖出"} and price <= 0:
            raise _error(TRADES_SHEET, row_number, "成交价", "必须大于 0")
        if kind == "分红" and cash_amount <= 0:
            raise _error(TRADES_SHEET, row_number, "现金金额", "必须大于 0")
        for field, value in (("手续费", commission), ("印花税", stamp_tax)):
            if value < 0:
                raise _error(TRADES_SHEET, row_number, field, "不能小于 0")

        note_value = values[columns["备注"]]
        trades.append(
            Trade(
                trade_date=_date(values[columns["日期"]], row_number),
                kind=cast(TradeKind, kind),
                grid_id=grid_id,
                quantity=quantity,
                price=price,
                commission=commission,
                stamp_tax=stamp_tax,
                cash_amount=cash_amount,
                note="" if note_value is None else str(note_value),
                excel_row=row_number,
            )
        )

    if not any(trade.kind == "买入" for trade in trades):
        raise WorkbookValidationError(f"{TRADES_SHEET}：至少需要一笔买入记录")
    return tuple(sorted(trades, key=lambda trade: (trade.trade_date, trade.excel_row)))


def load_workbook(path: str | Path) -> WorkbookData:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise WorkbookValidationError(f"找不到交易文件：{workbook_path}")
    try:
        book = openpyxl.load_workbook(workbook_path, data_only=True)
    except (OSError, ValueError, KeyError) as exc:
        raise WorkbookValidationError(f"无法读取交易文件：{exc}") from exc
    for sheet_name in (CONFIG_SHEET, TRADES_SHEET):
        if sheet_name not in book.sheetnames:
            raise WorkbookValidationError(f"缺少工作表：{sheet_name}")
    return WorkbookData(
        settings=_read_settings(book[CONFIG_SHEET]),
        trades=_read_trades(book[TRADES_SHEET]),
    )
