from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


def create_template(path: str | Path) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    book = openpyxl.Workbook()
    config = book.active
    config.title = "配置"
    config.append(["字段", "值", "说明"])
    config.append(["股票代码", "000001", "六位 A 股代码，建议单元格使用文本格式"])
    config.append(["图表标题", "网格波段 提款足迹", "显示在主图顶部"])
    config.append(["行情复权", "不复权", "首版固定填写：不复权"])
    config.append(["策略预算", 100000, "用于计算单品仓位"])

    trades = book.create_sheet("交易流水")
    trades.append(
        [
            "日期",
            "类型",
            "网格编号",
            "数量",
            "成交价",
            "手续费",
            "印花税",
            "现金金额",
            "备注",
        ]
    )
    trades.append([date(2025, 1, 2), "买入", "G3", 100, 11.20, 5, 0, 0, "示例：第三格买入"])
    trades.append([date(2025, 2, 10), "卖出", "G3", 100, 12.10, 5, 1.21, 0, "示例：第三格止盈"])
    trades.append([date(2025, 3, 3), "买入", "G2", 100, 11.50, 5, 0, 0, "示例：第二格持仓"])
    trades.append([date(2025, 6, 12), "分红", "", 0, 0, 0, 0, 50, "示例：现金分红"])

    dark_fill = PatternFill("solid", fgColor="111318")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="7A818D"))
    for sheet in (config, trades):
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = dark_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        sheet.auto_filter.ref = sheet.dimensions

    config.column_dimensions["A"].width = 16
    config.column_dimensions["B"].width = 24
    config.column_dimensions["C"].width = 46
    config["B2"].number_format = "@"

    widths = {
        "A": 14,
        "B": 10,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 14,
        "I": 30,
    }
    for column, width in widths.items():
        trades.column_dimensions[column].width = width
    for row in range(2, 1001):
        trades.cell(row, 1).number_format = "yyyy-mm-dd"
        for column in range(4, 9):
            trades.cell(row, column).number_format = "0.00"

    kind_validation = DataValidation(
        type="list", formula1='"买入,卖出,分红"', allow_blank=False
    )
    kind_validation.error = "类型只允许：买入、卖出、分红"
    kind_validation.errorTitle = "交易类型错误"
    trades.add_data_validation(kind_validation)
    kind_validation.add("B2:B1000")
    trades.conditional_formatting.add(
        "A2:I1000",
        FormulaRule(formula=['$B2="买入"'], font=Font(color="D90000")),
    )
    trades.conditional_formatting.add(
        "A2:I1000",
        FormulaRule(formula=['$B2="卖出"'], font=Font(color="008A3B")),
    )

    book.save(target)


if __name__ == "__main__":
    create_template(Path(__file__).resolve().parents[1] / "data" / "000001-平安银行.xlsx")
